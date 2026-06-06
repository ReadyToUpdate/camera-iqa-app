from __future__ import annotations

import math

import numpy as np
from openpyxl import load_workbook

from camera_iqa.stripe_fov import (
    StripeFovResult,
    analyze_stripe_fov_image,
    calculate_horizontal_fov_degrees,
    parse_sequence_from_filename,
    parse_zoom_from_filename,
    sort_results_by_sequence,
    write_stripe_fov_excel,
)


def make_vertical_stripes(
    *,
    image_width: int = 640,
    image_height: int = 360,
    chart_x: int = 120,
    chart_width: int = 360,
    stripe_width: int = 24,
) -> np.ndarray:
    image = np.full((image_height, image_width, 3), 128, dtype=np.uint8)
    for x in range(chart_x, chart_x + chart_width):
        stripe_index = (x - chart_x) // stripe_width
        value = 20 if stripe_index % 2 == 0 else 235
        image[:, x] = value
    return image


def test_analyze_stripe_fov_image_measures_black_stripes_when_chart_does_not_fill_frame() -> None:
    image = make_vertical_stripes()

    result = analyze_stripe_fov_image(image, distance_m=5.0, physical_black_width_m=0.01)

    assert result.image_width == 640
    assert result.image_height == 360
    assert result.chart_x == 120
    assert result.chart_width == 360
    assert result.black_stripe_count == 8
    assert result.black_stripe_px_mean == 24.0
    assert result.black_stripe_px_std == 0.0


def test_calculate_horizontal_fov_degrees_uses_image_width_distance_and_physical_width() -> None:
    fov = calculate_horizontal_fov_degrees(
        image_width_px=640,
        stripe_width_px=24.0,
        distance_m=5.0,
        physical_black_width_m=0.01,
    )

    expected_scene_width_m = 640 / 24.0 * 0.01
    expected = math.degrees(2 * math.atan(expected_scene_width_m / (2 * 5.0)))
    assert fov == expected


def test_analyze_stripe_fov_image_calculates_horizontal_fov() -> None:
    image = make_vertical_stripes(image_width=2560, image_height=1440, chart_x=380, chart_width=1200, stripe_width=40)

    result = analyze_stripe_fov_image(image, distance_m=8.0, physical_black_width_m=0.01)

    expected_scene_width_m = 2560 / 40.0 * 0.01
    expected = math.degrees(2 * math.atan(expected_scene_width_m / (2 * 8.0)))
    assert result.black_stripe_px_mean == 40.0
    assert result.horizontal_fov_deg == expected


def test_parse_zoom_from_filename_accepts_common_zoom_names() -> None:
    assert parse_zoom_from_filename("zoom_1.0.jpg") == 1.0
    assert parse_zoom_from_filename("camera-2.5x-distance5m.png") == 2.5
    assert parse_zoom_from_filename("倍率3.2_001.jpeg") == 3.2


def test_parse_sequence_from_filename_accepts_numeric_test_names() -> None:
    assert parse_sequence_from_filename("1.jpg") == 1
    assert parse_sequence_from_filename("090.png") == 90
    assert parse_sequence_from_filename("zoom_1.2_15.jpeg") == 15


def make_result(image: str, fov: float) -> StripeFovResult:
    return StripeFovResult(
        image=image,
        sequence=parse_sequence_from_filename(image),
        zoom=None,
        distance_m=5.0,
        image_width=2560,
        image_height=1440,
        chart_x=100,
        chart_width=1200,
        black_stripe_px_mean=40.0,
        black_stripe_px_std=0.5,
        black_stripe_count=20,
        horizontal_fov_deg=fov,
    )


def test_sort_results_by_sequence_uses_numeric_order() -> None:
    rows = [make_result("10.jpg", 3.0), make_result("2.jpg", 2.0), make_result("1.jpg", 1.0)]

    sorted_rows = sort_results_by_sequence(rows)

    assert [row.image for row in sorted_rows] == ["1.jpg", "2.jpg", "10.jpg"]


def test_write_stripe_fov_excel_outputs_sequence_and_fov(tmp_path) -> None:
    output = tmp_path / "fov.xlsx"
    rows = [make_result("2.jpg", 2.5), make_result("1.jpg", 3.5)]

    write_stripe_fov_excel(rows, output)

    workbook = load_workbook(output)
    sheet = workbook["FOV"]
    assert [sheet.cell(1, column).value for column in range(1, 4)] == ["序号", "水平视场角(度)", "图片"]
    assert sheet.cell(2, 1).value == 1
    assert sheet.cell(2, 2).value == 3.5
    assert sheet.cell(3, 1).value == 2
    assert sheet.cell(3, 2).value == 2.5
