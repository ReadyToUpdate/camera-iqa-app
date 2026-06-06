from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from camera_iqa.metrics import read_image, save_image, to_gray
from camera_iqa.pipeline import list_images


@dataclass(slots=True)
class StripeFovResult:
    image: str
    sequence: int | None
    zoom: float | None
    distance_m: float
    image_width: int
    image_height: int
    chart_x: int
    chart_width: int
    black_stripe_px_mean: float
    black_stripe_px_std: float
    black_stripe_count: int
    horizontal_fov_deg: float
    annotation_path: Path | None = None


def analyze_stripe_fov_path(
    path: str | Path,
    *,
    distance_m: float,
    physical_black_width_m: float = 0.01,
    zoom: float | None = None,
    annotation_path: str | Path | None = None,
) -> StripeFovResult:
    image_path = Path(path)
    result = analyze_stripe_fov_image(
        read_image(image_path),
        distance_m=distance_m,
        physical_black_width_m=physical_black_width_m,
        zoom=zoom if zoom is not None else parse_zoom_from_filename(image_path.name),
        sequence=parse_sequence_from_filename(image_path.name),
        image_name=image_path.name,
        annotation_path=annotation_path,
    )
    return result


def analyze_stripe_fov_image(
    image: np.ndarray,
    *,
    distance_m: float,
    physical_black_width_m: float = 0.01,
    zoom: float | None = None,
    sequence: int | None = None,
    image_name: str = "",
    annotation_path: str | Path | None = None,
) -> StripeFovResult:
    if distance_m <= 0:
        raise ValueError("distance_m must be positive")
    if physical_black_width_m <= 0:
        raise ValueError("physical_black_width_m must be positive")

    gray = to_gray(image) if image.ndim == 3 else image
    image_height, image_width = gray.shape
    chart_x, chart_width = locate_vertical_stripe_chart(gray)
    widths = measure_black_stripe_widths(gray, chart_x, chart_width)
    if not widths:
        raise ValueError("Cannot locate enough black stripes")

    mean_width = float(np.mean(widths))
    std_width = float(np.std(widths))
    horizontal_fov = calculate_horizontal_fov_degrees(
        image_width_px=image_width,
        stripe_width_px=mean_width,
        distance_m=distance_m,
        physical_black_width_m=physical_black_width_m,
    )

    output_path = Path(annotation_path) if annotation_path is not None else None
    if output_path is not None:
        save_image(output_path, make_annotation(image, chart_x, chart_width, widths, mean_width))

    return StripeFovResult(
        image=image_name,
        sequence=sequence,
        zoom=zoom,
        distance_m=distance_m,
        image_width=image_width,
        image_height=image_height,
        chart_x=chart_x,
        chart_width=chart_width,
        black_stripe_px_mean=mean_width,
        black_stripe_px_std=std_width,
        black_stripe_count=len(widths),
        horizontal_fov_deg=horizontal_fov,
        annotation_path=output_path,
    )


def locate_vertical_stripe_chart(gray: np.ndarray) -> tuple[int, int]:
    h, w = gray.shape
    y1 = h // 5
    y2 = h - y1
    sample = gray[y1:y2, :].astype(np.float32)
    profile = np.mean(sample, axis=0)
    median = float(np.median(profile))
    deviation = np.abs(profile - median)
    threshold = max(12.0, float(np.percentile(deviation, 75)) * 0.55)
    mask = deviation >= threshold

    close_size = max(3, w // 120)
    mask_u8 = (mask.astype(np.uint8) * 255)[None, :]
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (close_size, 1))
    mask_u8 = cv2.morphologyEx(mask_u8, cv2.MORPH_CLOSE, kernel)
    runs = boolean_runs(mask_u8[0] > 0)
    candidates = [(start, end) for start, end in runs if end - start >= max(20, w // 20)]
    if not candidates:
        raise ValueError("Cannot locate vertical stripe chart")

    start, end = max(candidates, key=lambda item: item[1] - item[0])
    return int(start), int(end - start)


def measure_black_stripe_widths(gray: np.ndarray, chart_x: int, chart_width: int) -> list[float]:
    h, _ = gray.shape
    y1 = h // 5
    y2 = h - y1
    profile = np.mean(gray[y1:y2, chart_x : chart_x + chart_width].astype(np.float32), axis=0)
    threshold, _ = cv2.threshold(profile.astype(np.uint8), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    black = profile <= float(threshold)
    runs = boolean_runs(black)
    widths = np.array([end - start for start, end in runs], dtype=np.float32)
    if widths.size == 0:
        return []

    median_width = float(np.median(widths))
    if median_width <= 0:
        return []
    filtered = widths[(widths >= median_width * 0.65) & (widths <= median_width * 1.35)]
    if filtered.size == 0:
        return []
    return [float(width) for width in filtered]


def boolean_runs(mask: np.ndarray) -> list[tuple[int, int]]:
    values = np.asarray(mask, dtype=bool)
    if values.size == 0:
        return []
    padded = np.concatenate(([False], values, [False]))
    changes = np.flatnonzero(padded[1:] != padded[:-1])
    return [(int(changes[index]), int(changes[index + 1])) for index in range(0, len(changes), 2)]


def calculate_horizontal_fov_degrees(
    *,
    image_width_px: int,
    stripe_width_px: float,
    distance_m: float,
    physical_black_width_m: float,
) -> float:
    if image_width_px <= 0:
        raise ValueError("image_width_px must be positive")
    if stripe_width_px <= 0:
        raise ValueError("stripe_width_px must be positive")
    if distance_m <= 0:
        raise ValueError("distance_m must be positive")
    if physical_black_width_m <= 0:
        raise ValueError("physical_black_width_m must be positive")

    scene_width_m = image_width_px / stripe_width_px * physical_black_width_m
    return float(math.degrees(2 * math.atan(scene_width_m / (2 * distance_m))))


def parse_zoom_from_filename(filename: str) -> float | None:
    patterns = (
        r"(?:zoom|倍率)[_-]?(\d+(?:\.\d+)?)",
        r"(\d+(?:\.\d+)?)\s*x",
    )
    for pattern in patterns:
        match = re.search(pattern, filename, flags=re.IGNORECASE)
        if match:
            return float(match.group(1))
    return None


def parse_sequence_from_filename(filename: str) -> int | None:
    stem = Path(filename).stem
    if stem.isdigit():
        return int(stem)
    matches = re.findall(r"\d+", stem)
    if not matches:
        return None
    return int(matches[-1])


def read_distance_csv(path: str | Path) -> dict[str, float]:
    distances: dict[str, float] = {}
    with Path(path).open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or "image" not in reader.fieldnames or "distance_m" not in reader.fieldnames:
            raise ValueError("Distance CSV must contain image,distance_m columns")
        for row in reader:
            image_name = (row.get("image") or "").strip()
            distance_text = (row.get("distance_m") or "").strip()
            if not image_name or not distance_text:
                continue
            distances[image_name] = float(distance_text)
    return distances


def analyze_stripe_fov_folder(
    input_dir: str | Path,
    *,
    distances: dict[str, float],
    physical_black_width_m: float = 0.01,
    annotation_dir: str | Path | None = None,
) -> list[StripeFovResult]:
    output_dir = Path(annotation_dir) if annotation_dir is not None else None
    rows: list[StripeFovResult] = []
    for image_path in list_images(input_dir):
        distance_m = distances.get(image_path.name)
        if distance_m is None:
            raise ValueError(f"Missing distance_m for {image_path.name}")
        annotation_path = None
        if output_dir is not None:
            annotation_path = output_dir / f"{image_path.stem}_stripe_fov.jpg"
        rows.append(
            analyze_stripe_fov_path(
                image_path,
                distance_m=distance_m,
                physical_black_width_m=physical_black_width_m,
                annotation_path=annotation_path,
            )
        )
    return rows


def sort_results_by_sequence(rows: list[StripeFovResult]) -> list[StripeFovResult]:
    return sorted(rows, key=lambda row: (row.sequence is None, row.sequence or 0, row.image))


def write_stripe_fov_csv(rows: list[StripeFovResult], output: str | Path) -> None:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "sequence",
        "image",
        "zoom",
        "distance_m",
        "image_width",
        "image_height",
        "chart_x",
        "chart_width",
        "black_stripe_px_mean",
        "black_stripe_px_std",
        "black_stripe_count",
        "horizontal_fov_deg",
        "annotation_path",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in sort_results_by_sequence(rows):
            writer.writerow(
                {
                    "sequence": "" if row.sequence is None else row.sequence,
                    "image": row.image,
                    "zoom": "" if row.zoom is None else f"{row.zoom:.3f}",
                    "distance_m": f"{row.distance_m:.4f}",
                    "image_width": row.image_width,
                    "image_height": row.image_height,
                    "chart_x": row.chart_x,
                    "chart_width": row.chart_width,
                    "black_stripe_px_mean": f"{row.black_stripe_px_mean:.3f}",
                    "black_stripe_px_std": f"{row.black_stripe_px_std:.3f}",
                    "black_stripe_count": row.black_stripe_count,
                    "horizontal_fov_deg": f"{row.horizontal_fov_deg:.6f}",
                    "annotation_path": "" if row.annotation_path is None else str(row.annotation_path),
                }
            )


def write_stripe_fov_excel(rows: list[StripeFovResult], output: str | Path) -> Path:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FOV"
    headers = [
        "序号",
        "水平视场角(度)",
        "图片",
        "距离(m)",
        "黑条平均宽度(px)",
        "黑条宽度标准差(px)",
        "黑条数量",
        "图像宽度(px)",
        "图像高度(px)",
        "测试卡起点x(px)",
        "测试卡宽度(px)",
        "倍率",
        "标注图",
    ]
    sheet.append(headers)
    for row in sort_results_by_sequence(rows):
        sheet.append(
            [
                row.sequence,
                round(row.horizontal_fov_deg, 6),
                row.image,
                round(row.distance_m, 4),
                round(row.black_stripe_px_mean, 3),
                round(row.black_stripe_px_std, 3),
                row.black_stripe_count,
                row.image_width,
                row.image_height,
                row.chart_x,
                row.chart_width,
                None if row.zoom is None else round(row.zoom, 3),
                "" if row.annotation_path is None else str(row.annotation_path),
            ]
        )

    header_fill = PatternFill("solid", fgColor="243447")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")
    for column in range(1, sheet.max_column + 1):
        width = 16
        if column in (3, 13):
            width = 34
        sheet.column_dimensions[get_column_letter(column)].width = width

    workbook.save(output_path)
    return output_path


def make_annotation(
    image: np.ndarray,
    chart_x: int,
    chart_width: int,
    widths: list[float],
    mean_width: float,
) -> np.ndarray:
    annotated = image.copy()
    if annotated.ndim == 2:
        annotated = cv2.cvtColor(annotated, cv2.COLOR_GRAY2BGR)
    h, _ = annotated.shape[:2]
    cv2.rectangle(annotated, (chart_x, 0), (chart_x + chart_width, h - 1), (0, 220, 0), 2)
    text = f"black stripes: n={len(widths)} mean={mean_width:.2f}px"
    cv2.putText(annotated, text, (12, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 220, 0), 2, cv2.LINE_AA)
    return annotated
