from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from camera_iqa.catalog import all_metric_keys, metric_group_for
from camera_iqa.models import ImageResult


METRIC_COLUMNS = all_metric_keys()


def export_excel(results: list[ImageResult], output_path: str | Path, config: dict[str, Any], input_folder: str | Path | None = None) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    ws = workbook.active
    ws.title = "Dashboard"
    build_dashboard(ws, results, input_folder)
    build_details(workbook.create_sheet("Details"), results)
    build_samples(workbook.create_sheet("Defect Samples"), results)
    build_config(workbook.create_sheet("Config"), config)

    workbook.save(output)
    return output


def build_dashboard(ws, results: list[ImageResult], input_folder: str | Path | None) -> None:
    total = len(results)
    done = sum(1 for result in results if result.status == "done")
    passed = sum(1 for result in results if result.verdict == "pass")
    warn = sum(1 for result in results if result.severity == "warn")
    fail = sum(1 for result in results if result.severity == "fail")

    ws["A1"] = "安防摄像机图像客观测试报告"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A3"] = "输入文件夹"
    ws["B3"] = str(input_folder or "")
    ws["A5"] = "总图片数"
    ws["B5"] = total
    ws["C5"] = "完成"
    ws["D5"] = done
    ws["E5"] = "通过"
    ws["F5"] = passed
    ws["A6"] = "警告"
    ws["B6"] = warn
    ws["C6"] = "失败"
    ws["D6"] = fail
    ws["E6"] = "通过率"
    ws["F6"] = f"{(passed / total * 100):.1f}%" if total else "0.0%"

    defect_counts = Counter(defect.code for result in results for defect in result.defects)
    ws["A9"] = "缺陷分布"
    ws["A9"].font = Font(bold=True)
    ws.append(["缺陷", "数量"])
    for code, count in defect_counts.most_common():
        ws.append([code, count])

    metric_start = max(13, ws.max_row + 2)
    ws.cell(metric_start, 1, "关键指标统计").font = Font(bold=True)
    ws.cell(metric_start + 1, 1, "指标")
    ws.cell(metric_start + 1, 2, "类别")
    ws.cell(metric_start + 1, 3, "平均值")
    ws.cell(metric_start + 1, 4, "最小值")
    ws.cell(metric_start + 1, 5, "最大值")
    done_results = [result for result in results if result.status == "done"]
    for row_index, metric in enumerate(METRIC_COLUMNS, start=metric_start + 2):
        values = [result.metrics.get(metric, 0.0) for result in done_results]
        ws.cell(row_index, 1, metric)
        ws.cell(row_index, 2, metric_group_for(metric))
        ws.cell(row_index, 3, round(sum(values) / len(values), 4) if values else "")
        ws.cell(row_index, 4, round(min(values), 4) if values else "")
        ws.cell(row_index, 5, round(max(values), 4) if values else "")

    style_sheet(ws)


def build_details(ws, results: list[ImageResult]) -> None:
    headers = ["file", "path", "status", "verdict", "severity", "width", "height", "defects", "error"] + METRIC_COLUMNS
    ws.append(headers)
    for result in results:
        row = result.as_row()
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws, freeze="A2")


def build_samples(ws, results: list[ImageResult]) -> None:
    ws.append(["file", "category", "severity", "defects", "preview"])
    row = 2
    for result in results:
        if not result.defects:
            continue
        ws.cell(row, 1, result.path.name)
        ws.cell(row, 2, "、".join(sorted({defect.category for defect in result.defects})))
        ws.cell(row, 3, result.severity)
        ws.cell(row, 4, result.defect_codes)
        if result.overlay_path and result.overlay_path.exists():
            try:
                image = ExcelImage(str(result.overlay_path))
                image.width = 220
                image.height = 140
                ws.add_image(image, f"E{row}")
                ws.row_dimensions[row].height = 108
            except Exception:
                ws.cell(row, 5, str(result.overlay_path))
        row += 1
    style_sheet(ws, freeze="A2")


def build_config(ws, config: dict[str, Any]) -> None:
    ws.append(["section", "key", "value"])
    for key, value in config.get("metadata", {}).items():
        ws.append(["metadata", key, str(value)])
    ws.append(["runtime", "source_path", config.get("_source_path", "")])
    for metric, rules in config.get("thresholds", {}).items():
        for rule_key, value in rules.items():
            ws.append([metric, rule_key, value])
    style_sheet(ws, freeze="A2")


def style_sheet(ws, freeze: str | None = None) -> None:
    if freeze:
        ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="243447")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for column in range(1, min(ws.max_column, 12) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18
    ws.column_dimensions["B"].width = 32
